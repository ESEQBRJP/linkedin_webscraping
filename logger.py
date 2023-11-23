import json
import openpyxl
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter



class XlsxUteis:

    def __init__(self):
        self.filename = f"Report_.xlsx"
        self.path = './out'
        self.full_path = f'{self.path}/{self.filename}'
        self.workbook = openpyxl.Workbook()
        self.name_sheet = None

    def append_sheet(self, name: str):
        self.name_sheet = self.workbook.create_sheet(name, -1)

    def append_header_sheet(self, header: list):
        for x in range(1, len(header)+1):
            self.name_sheet.cell(row=1, column=x, value=header[x-1])

    def append_data_sheet(self,  data: list):

        linha = len(data)
        coluna = len(data[0])
        matriz = (linha, coluna,)

        for x in range(1, matriz[0]+1):
            for y in range(1, matriz[1]+1):
                self.name_sheet.cell(row=x+1, column=y, value=data[x-1][y-1])

    def save_workbook(self):
        self.workbook.save(self.full_path)

    def _backgroud_style_colunm_full(self):

        patternFill = PatternFill(start_color="00FF8080", end_color="00FF8080", fill_type="solid")
        ft = Font(color="FF0000")
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))


        for position_colunm, coluns in enumerate(self.name_sheet.iter_cols(min_col=1, max_row=1)):
            for cell in coluns:
                if cell.value == 'DIFF':
                    for position_row, row in enumerate(self.name_sheet.iter_rows(min_row=2)):
                        if row[position_colunm].value == 'DIFF':
                            print(row[5].value)
                            for i in range(1, position_colunm+2):
                                self.name_sheet.cell(position_row + 2, i).fill = patternFill
                                self.name_sheet.cell(position_row + 2, i).border = thin_border

    def _background_default_style(self):
        patternHeader = PatternFill(start_color="00808080", end_color="00808080", fill_type="solid")
        patternGeral = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
        alignment = Alignment(horizontal='center', vertical=None, textRotation=0, wrapText=None, shrinkToFit=None, indent=0,
                  relativeIndent=0, justifyLastLine=None, readingOrder=0, text_rotation=None, wrap_text=None,
                  shrink_to_fit=None, mergeCell=None)

        ft = Font(color="FF0000")
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for position_colunm, coluns in enumerate(self.name_sheet.iter_cols(min_col=1, max_row=1)):
            self.name_sheet.cell(1, position_colunm + 1).fill = patternHeader
            self.name_sheet.cell(1, position_colunm + 1).border = thin_border
            self.name_sheet.column_dimensions[get_column_letter(position_colunm+1)].auto_size = True
            for cell in coluns:
                letter_size = len(str(cell.value))
            for position_row, row in enumerate(self.name_sheet.iter_rows(min_row=2)):
                if row[position_colunm].value == '-':
                    for i in range(1, position_colunm+2):
                        self.name_sheet.cell(position_row + 2, i).fill = patternHeader
                        self.name_sheet.cell(position_row + 2, i).border = thin_border
                        self.name_sheet.cell(1, position_colunm + 1).alignment = alignment
                        self.name_sheet.column_dimensions[get_column_letter(position_colunm + 1)].width = 20

                else:
                    for i in range(1, position_colunm+2):
                        self.name_sheet.cell(position_row + 2, i).fill = patternGeral
                        self.name_sheet.cell(position_row + 2, i).border = thin_border
                        self.name_sheet.cell(1, position_colunm + 1).alignment = alignment
                        self.name_sheet.column_dimensions[get_column_letter(position_colunm + 1)].width = 20

    def append_sheet_log_data(self, name: str, header, data):
        self.append_sheet(name)
        self.append_header_sheet(header)
        self.append_data_sheet(data)
        self._background_default_style()
        self._backgroud_style_colunm_full()
        self.save_workbook()